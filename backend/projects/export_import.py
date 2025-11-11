"""
Export/Import utilities for projects
"""
import io
import csv
from datetime import datetime
from django.http import HttpResponse
from django.db import transaction
from rest_framework import status
from rest_framework.response import Response
from .models import Project, ProjectGroup
from students.models import Student
from advisors.models import Advisor
from majors.models import Major


def export_projects_to_csv(queryset, filename=None):
    """
    Export projects to CSV format
    """
    if filename is None:
        filename = f'projects_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Project ID', 'Topic (Lao)', 'Topic (English)', 'Advisor Name',
        'Student IDs', 'Student Names', 'Status', 'Defense Date',
        'Defense Time', 'Defense Room', 'Final Grade', 'Final Score',
        'Created At', 'Updated At'
    ])
    
    # Write data
    for project in queryset:
        # Get ProjectGroup for this project
        project_group = ProjectGroup.objects.filter(project_id=project.project_id).first()
        student_ids = []
        student_names = []
        topic_lao = ''
        topic_eng = ''
        advisor_name = ''
        status = project.status
        defense_date = None
        defense_time = None
        defense_room = ''
        final_grade = ''
        final_score = ''
        
        if project_group:
            topic_lao = project_group.topic_lao or ''
            topic_eng = project_group.topic_eng or ''
            advisor_name = project_group.advisor_name or ''
            status = project_group.status or project.status
            defense_date = project_group.defense_date
            defense_time = project_group.defense_time
            defense_room = project_group.defense_room or ''
            final_grade = project_group.final_grade or ''
            # Calculate final_score from committee scores
            scores = [s for s in [
                project_group.main_advisor_score,
                project_group.main_committee_score,
                project_group.second_committee_score,
                project_group.third_committee_score
            ] if s is not None]
            final_score = sum(scores) / len(scores) if scores else ''
            
            # Get students from ProjectGroup
            from projects.models import ProjectStudent
            project_students = ProjectStudent.objects.filter(project_group=project_group)
            for ps in project_students:
                student_ids.append(ps.student.student_id if hasattr(ps.student, 'student_id') else str(ps.student.id))
                student_names.append(f"{ps.student.get_full_name() if hasattr(ps.student, 'get_full_name') else ps.student.username}")
        
        writer.writerow([
            project.project_id,
            topic_lao,
            topic_eng,
            advisor_name,
            ', '.join(student_ids),
            ', '.join(student_names),
            status,
            defense_date.strftime('%Y-%m-%d') if defense_date else '',
            defense_time.strftime('%H:%M') if defense_time else '',
            defense_room,
            final_grade,
            str(final_score) if final_score else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
            project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else '',
        ])
    
    return response


def export_projects_to_excel(queryset, filename=None):
    """
    Export projects to Excel format using openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    if filename is None:
        filename = f'projects_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    # Headers
    headers = [
        'Project ID', 'Topic (Lao)', 'Topic (English)', 'Advisor Name',
        'Student IDs', 'Student Names', 'Status', 'Defense Date',
        'Defense Time', 'Defense Room', 'Final Grade', 'Final Score',
        'Created At', 'Updated At'
    ]
    
    # Style header row
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_num, project in enumerate(queryset, 2):
        # Get ProjectGroup for this project
        project_group = ProjectGroup.objects.filter(project_id=project.project_id).first()
        student_ids = []
        student_names = []
        topic_lao = ''
        topic_eng = ''
        advisor_name = ''
        status = project.status
        defense_date = None
        defense_time = None
        defense_room = ''
        final_grade = ''
        final_score = ''
        
        if project_group:
            topic_lao = project_group.topic_lao or ''
            topic_eng = project_group.topic_eng or ''
            advisor_name = project_group.advisor_name or ''
            status = project_group.status or project.status
            defense_date = project_group.defense_date
            defense_time = project_group.defense_time
            defense_room = project_group.defense_room or ''
            final_grade = project_group.final_grade or ''
            # Calculate final_score from committee scores
            scores = [s for s in [
                project_group.main_advisor_score,
                project_group.main_committee_score,
                project_group.second_committee_score,
                project_group.third_committee_score
            ] if s is not None]
            final_score = sum(scores) / len(scores) if scores else ''
            
            # Get students from ProjectGroup
            from projects.models import ProjectStudent
            project_students = ProjectStudent.objects.filter(project_group=project_group)
            for ps in project_students:
                student_ids.append(ps.student.student_id if hasattr(ps.student, 'student_id') else str(ps.student.id))
                student_names.append(f"{ps.student.get_full_name() if hasattr(ps.student, 'get_full_name') else ps.student.username}")
        
        ws.cell(row=row_num, column=1, value=project.project_id)
        ws.cell(row=row_num, column=2, value=topic_lao)
        ws.cell(row=row_num, column=3, value=topic_eng)
        ws.cell(row=row_num, column=4, value=advisor_name)
        ws.cell(row=row_num, column=5, value=', '.join(student_ids))
        ws.cell(row=row_num, column=6, value=', '.join(student_names))
        ws.cell(row=row_num, column=7, value=status)
        ws.cell(row=row_num, column=8, value=defense_date.strftime('%Y-%m-%d') if defense_date else '')
        ws.cell(row=row_num, column=9, value=defense_time.strftime('%H:%M') if defense_time else '')
        ws.cell(row=row_num, column=10, value=defense_room)
        ws.cell(row=row_num, column=11, value=final_grade)
        ws.cell(row=row_num, column=12, value=str(final_score) if final_score else '')
        ws.cell(row=row_num, column=13, value=project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '')
        ws.cell(row=row_num, column=14, value=project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def import_projects_from_csv(file, academic_year=None, user=None):
    """
    Import projects from CSV file
    Returns: (success_count, error_count, errors)
    """
    errors = []
    success_count = 0
    error_count = 0
    
    try:
        # Read CSV
        decoded_file = file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        
        with transaction.atomic():
            for row_num, row in enumerate(reader, start=2):  # Start at 2 (row 1 is header)
                try:
                    # Validate required fields
                    if not row.get('Project ID'):
                        errors.append(f"Row {row_num}: Project ID is required")
                        error_count += 1
                        continue
                    
                    project_id = row['Project ID'].strip()
                    
                    # Create or update ProjectGroup
                    project_group, created = ProjectGroup.objects.get_or_create(
                        project_id=project_id,
                        defaults={
                            'topic_lao': row.get('Topic (Lao)', '').strip(),
                            'topic_eng': row.get('Topic (English)', '').strip(),
                            'advisor_name': row.get('Advisor Name', '').strip(),
                            'status': row.get('Status', 'Pending').strip(),
                            'final_grade': row.get('Final Grade', '').strip() or None,
                        }
                    )
                    
                    if not created:
                        # Update existing project group
                        if row.get('Topic (Lao)'):
                            project_group.topic_lao = row['Topic (Lao)'].strip()
                        if row.get('Topic (English)'):
                            project_group.topic_eng = row['Topic (English)'].strip()
                        if row.get('Advisor Name'):
                            project_group.advisor_name = row['Advisor Name'].strip()
                        if row.get('Status'):
                            project_group.status = row['Status'].strip()
                        if row.get('Final Grade'):
                            project_group.final_grade = row['Final Grade'].strip() or None
                        project_group.save()
                    
                    # Create or update Project (legacy model)
                    project, _ = Project.objects.get_or_create(
                        project_id=project_id,
                        defaults={
                            'title': row.get('Topic (English)', '').strip() or row.get('Topic (Lao)', '').strip(),
                            'status': row.get('Status', 'Pending').strip(),
                        }
                    )
                    
                    if not _:
                        # Update existing project
                        if row.get('Topic (English)'):
                            project.title = row['Topic (English)'].strip()
                        if row.get('Status'):
                            project.status = row['Status'].strip()
                        project.save()
                    
                    # Handle students if provided
                    if row.get('Student IDs'):
                        student_ids = [s.strip() for s in row['Student IDs'].split(',')]
                        # This would require more complex logic to link students
                        # For now, just log that students were found
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
                    continue
        
        return success_count, error_count, errors
        
    except Exception as e:
        errors.append(f"File processing error: {str(e)}")
        return 0, 1, errors

